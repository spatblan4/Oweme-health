from __future__ import annotations

from csv import DictReader
from pathlib import Path
import re
from typing import Any

from worker.extract.pdf import extract_pdf_text
from worker.extract.text import extract_text


def _first_match(pattern: str, text: str) -> str | None:
    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
    if not match:
        return None
    value = match.group(1).strip()
    return value or None


def _parse_pdf_page(page_text: str) -> list[dict[str, Any]]:
    lines = [line.strip() for line in page_text.splitlines() if line.strip()]
    if not lines:
        return []

    text = "\n".join(lines)
    rows: list[dict[str, Any]] = []

    if re.search(r"receipt of payment|payment id|payment type|amount:", text, re.IGNORECASE):
        merchant = _first_match(r"^([A-Z0-9][^\n]+)$", text)
        if merchant and merchant.lower().startswith("page "):
            merchant = None
        rows.append(
            {
                "merchant": merchant
                or _first_match(r"received from:\s*([^\n]+)", text)
                or _first_match(r"facility\s*:\s*([^\n]+)", text)
                or _first_match(r"provider:\s*([^\n]+)", text),
                "amount": _first_match(r"amount:\s*([$()0-9.,-]+)", text)
                or _first_match(r"patient payment\s*([$()0-9.,-]+)", text),
                "raw_text": text,
            }
        )
        return rows

    if re.search(r"service date|provider:|facility\s*:", text, re.IGNORECASE):
        rows.append(
            {
                "provider": _first_match(r"provider:\s*([^\n]+?)(?=\s+(?:address|employer|insurances?|icd codes|billed|procedure codes|notes)\b|$)", text)
                or _first_match(r"facility\s*:\s*([^\n]+)", text)
                or _first_match(r"resource name:\s*([^\n]+)", text),
                "service_date": _first_match(r"service date\s*:?\s*([0-9/.-]+)", text),
                "raw_text": text,
            }
        )
        return rows

    return []


def _extract_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [dict(row) for row in DictReader(handle)]


def _extract_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        headers: list[str] | None = None
        for header_row in iterator:
            values = list(header_row)
            if not any(value is not None and value != "" for value in values):
                continue
            headers = [str(cell).strip() if cell is not None else "" for cell in values]
            break
        if headers is None:
            continue
        headers = [header or f"column_{index + 1}" for index, header in enumerate(headers)]
        for row in iterator:
            values = list(row)
            if not any(value is not None and value != "" for value in values):
                continue
            rows.append(
                {
                    headers[index]: value
                    for index, value in enumerate(values)
                    if index < len(headers)
                }
            )
    return rows


def _extract_xls_rows(path: Path) -> list[dict[str, Any]]:
    import xlrd

    workbook = xlrd.open_workbook(path)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        headers: list[str] | None = None
        start_row_index = 0
        for row_index in range(sheet.nrows):
            values = sheet.row_values(row_index)
            if not any(str(value).strip() for value in values):
                continue
            headers = [str(cell).strip() if cell is not None else "" for cell in values]
            start_row_index = row_index + 1
            break
        if headers is None:
            continue
        headers = [header or f"column_{index + 1}" for index, header in enumerate(headers)]
        for row_index in range(start_row_index, sheet.nrows):
            values = sheet.row_values(row_index)
            if not any(str(value).strip() for value in values):
                continue
            rows.append(
                {
                    headers[index]: value
                    for index, value in enumerate(values)
                    if index < len(headers)
                }
            )
    return rows


def extract_tables(path: Path, deps: dict | None = None) -> list[dict]:
    resolved = deps or {}
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return _extract_csv_rows(path)

    if suffix == ".xlsx":
        return _extract_xlsx_rows(path)

    if suffix == ".xls":
        return _extract_xls_rows(path)

    if suffix == ".pdf":
        raw_text = extract_pdf_text(path, deps=resolved)
        pages = re.split(r"^===== Page \d+ =====\s*$", raw_text, flags=re.MULTILINE)
        rows: list[dict] = []
        for page in pages:
            rows.extend(_parse_pdf_page(page))
        return rows

    text = extract_text(path, deps=resolved)
    if text:
        return _parse_pdf_page(text)

    return []
